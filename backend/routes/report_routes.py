"""
BudgetIQ – Report Export Routes (PDF & Excel)
"""
import io
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import Income, Expense, User
from auth import get_current_user

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/reports", tags=["Reports"])


def get_date_range(period: str):
    """Get start and end dates based on period (weekly or monthly)."""
    now = datetime.now(timezone.utc)
    if period == "weekly":
        start = now - timedelta(days=7)
    else:
        start = now.replace(day=1)
    return start, now


@router.get("/pdf")
def export_pdf(
    period: str = Query("monthly", regex="^(weekly|monthly)$"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Generate and download a PDF report."""
    start, end = get_date_range(period)

    # Fetch data
    incomes = db.query(Income).filter(
        Income.user_id == user.id, Income.date >= start, Income.date <= end
    ).order_by(Income.date).all()

    expenses = db.query(Expense).filter(
        Expense.user_id == user.id, Expense.date >= start, Expense.date <= end
    ).order_by(Expense.date).all()

    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)
    balance = total_income - total_expense

    # Build PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=20, textColor=colors.HexColor("#6C63FF"))
    elements.append(Paragraph("BudgetIQ Financial Report", title_style))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f"<b>Period:</b> {period.capitalize()} | {start.strftime('%d %b %Y')} – {end.strftime('%d %b %Y')}",
        styles['Normal']
    ))
    elements.append(Paragraph(f"<b>User:</b> {user.name} ({user.email})", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary table
    summary_data = [
        ["Total Income", "Total Expenses", "Balance"],
        [f"₹{total_income:,.2f}", f"₹{total_expense:,.2f}", f"₹{balance:,.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6C63FF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Income table
    elements.append(Paragraph("<b>Income Details</b>", styles['Heading2']))
    if incomes:
        inc_data = [["Date", "Source", "Amount"]]
        for i in incomes:
            inc_data.append([i.date.strftime("%d %b %Y"), i.source, f"₹{i.amount:,.2f}"])
        inc_table = Table(inc_data, colWidths=[2*inch, 2.5*inch, 1.5*inch])
        inc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(inc_table)
    else:
        elements.append(Paragraph("No income entries for this period.", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Expense table
    elements.append(Paragraph("<b>Expense Details</b>", styles['Heading2']))
    if expenses:
        exp_data = [["Date", "Category", "Description", "Amount"]]
        for e in expenses:
            exp_data.append([
                e.date.strftime("%d %b %Y"),
                e.category,
                (e.description or "")[:30],
                f"₹{e.amount:,.2f}"
            ])
        exp_table = Table(exp_data, colWidths=[1.3*inch, 1.5*inch, 2*inch, 1.2*inch])
        exp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#FF5252")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF0F0")]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(exp_table)
    else:
        elements.append(Paragraph("No expense entries for this period.", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)

    filename = f"BudgetIQ_{period}_report_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/excel")
def export_excel(
    period: str = Query("monthly", regex="^(weekly|monthly)$"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """Generate and download an Excel report."""
    start, end = get_date_range(period)

    incomes = db.query(Income).filter(
        Income.user_id == user.id, Income.date >= start, Income.date <= end
    ).order_by(Income.date).all()

    expenses = db.query(Expense).filter(
        Expense.user_id == user.id, Expense.date >= start, Expense.date <= end
    ).order_by(Expense.date).all()

    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)
    balance = total_income - total_expense

    wb = Workbook()
    # Styling
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='6C63FF', fill_type='solid')
    green_fill = PatternFill(start_color='4CAF50', fill_type='solid')
    red_fill = PatternFill(start_color='FF5252', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ─── Summary Sheet ───
    ws = wb.active
    ws.title = "Summary"
    ws.merge_cells('A1:C1')
    ws['A1'] = "BudgetIQ Financial Report"
    ws['A1'].font = Font(name='Calibri', bold=True, size=16, color='6C63FF')
    ws['A2'] = f"Period: {period.capitalize()} | {start.strftime('%d %b %Y')} – {end.strftime('%d %b %Y')}"
    ws['A3'] = f"User: {user.name}"

    for col, header in enumerate(["Total Income", "Total Expenses", "Balance"], 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    ws.cell(row=6, column=1, value=total_income).border = border
    ws.cell(row=6, column=2, value=total_expense).border = border
    ws.cell(row=6, column=3, value=balance).border = border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # ─── Income Sheet ───
    ws_inc = wb.create_sheet("Income")
    inc_headers = ["Date", "Source", "Amount"]
    for col, h in enumerate(inc_headers, 1):
        cell = ws_inc.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for row, i in enumerate(incomes, 2):
        ws_inc.cell(row=row, column=1, value=i.date.strftime("%d %b %Y")).border = border
        ws_inc.cell(row=row, column=2, value=i.source).border = border
        ws_inc.cell(row=row, column=3, value=i.amount).border = border
    ws_inc.column_dimensions['A'].width = 15
    ws_inc.column_dimensions['B'].width = 25
    ws_inc.column_dimensions['C'].width = 15

    # ─── Expense Sheet ───
    ws_exp = wb.create_sheet("Expenses")
    exp_headers = ["Date", "Category", "Description", "Amount"]
    for col, h in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = red_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    for row, e in enumerate(expenses, 2):
        ws_exp.cell(row=row, column=1, value=e.date.strftime("%d %b %Y")).border = border
        ws_exp.cell(row=row, column=2, value=e.category).border = border
        ws_exp.cell(row=row, column=3, value=e.description or "").border = border
        ws_exp.cell(row=row, column=4, value=e.amount).border = border
    ws_exp.column_dimensions['A'].width = 15
    ws_exp.column_dimensions['B'].width = 20
    ws_exp.column_dimensions['C'].width = 30
    ws_exp.column_dimensions['D'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"BudgetIQ_{period}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
